"""
DataSecurity Daily Report - GitHub Actions / Google Drive API Version
---------------------------------------------------------------------
Runs entirely in the cloud via GitHub Actions.
No local PC, no G:\ drive, no user login required.

- Downloads CSV files from Google Drive DataSecurity folder via API
  (uses service account for reading)
- Consolidates into a single formatted Excel report
- Uploads report to Consolidated_Report folder using OAuth
  (uses your personal Google account for writing - no quota issues)
- Saves a permanent archive copy to Reports_Archive folder
- Zapier detects new file in Consolidated_Report and sends email

Schedule: Daily at 7:30 AM via GitHub Actions cron
"""

import os
import re
import json
import logging
import tempfile
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from google.oauth2 import service_account
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload


# ============================================================
# CONFIGURATION
# ============================================================

GDRIVE_SOURCE_FOLDER_NAME  = "DataSecurity"        # folder with CSV files
GDRIVE_REPORTS_FOLDER_NAME = "Consolidated_Report" # Zapier watches this
GDRIVE_ARCHIVE_FOLDER_NAME = "Reports_Archive"     # permanent copy folder

SCOPES = ["https://www.googleapis.com/auth/drive"]


# ============================================================
# LOGGING
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)


# ============================================================
# GOOGLE DRIVE API CONNECTIONS
# Two connections:
#   1. Service account — for READING CSV files from DataSecurity folder
#   2. OAuth (your account) — for WRITING reports to output folders
# ============================================================

def get_service_account_service():
    """Service account connection — used for downloading CSVs."""
    creds_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not creds_json:
        raise ValueError("GOOGLE_SERVICE_ACCOUNT_JSON secret not set.")

    creds_dict  = json.loads(creds_json)
    credentials = service_account.Credentials.from_service_account_info(
        creds_dict, scopes=SCOPES
    )
    service = build("drive", "v3", credentials=credentials)
    logger.info("Connected to Google Drive API (service account)")
    return service


def get_oauth_service():
    """
    OAuth connection using your personal Google account.
    Used for uploading reports — no storage quota issues.
    Credentials stored as GitHub secrets.
    """
    client_id     = os.environ.get("GOOGLE_OAUTH_CLIENT_ID")
    client_secret = os.environ.get("GOOGLE_OAUTH_CLIENT_SECRET")
    refresh_token = os.environ.get("GOOGLE_OAUTH_REFRESH_TOKEN")

    if not all([client_id, client_secret, refresh_token]):
        raise ValueError(
            "Missing OAuth secrets. Ensure GOOGLE_OAUTH_CLIENT_ID, "
            "GOOGLE_OAUTH_CLIENT_SECRET and GOOGLE_OAUTH_REFRESH_TOKEN "
            "are set as GitHub secrets."
        )

    creds = Credentials(
        token=None,
        refresh_token=refresh_token,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=client_id,
        client_secret=client_secret,
        scopes=SCOPES
    )

    # Refresh to get a valid access token
    creds.refresh(Request())
    service = build("drive", "v3", credentials=creds)
    logger.info("Connected to Google Drive API (OAuth - your account)")
    return service


# ============================================================
# FOLDER HELPERS
# ============================================================

def get_folder_id(service, folder_name, parent_id=None):
    """Find a Google Drive folder by name and return its ID."""
    query = (
        f"name='{folder_name}' "
        f"and mimeType='application/vnd.google-apps.folder' "
        f"and trashed=false"
    )
    if parent_id:
        query += f" and '{parent_id}' in parents"

    results = service.files().list(
        q=query,
        fields="files(id, name)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True
    ).execute()

    files = results.get("files", [])
    if not files:
        logger.warning(f"Folder '{folder_name}' not found in Google Drive")
        return None

    logger.info(f"Found folder '{folder_name}' (ID: {files[0]['id']})")
    return files[0]["id"]


def get_or_create_folder(service, folder_name, parent_id=None):
    """Get folder ID, creating it if it doesn't exist."""
    folder_id = get_folder_id(service, folder_name, parent_id)
    if not folder_id:
        logger.info(f"Creating folder: '{folder_name}'")
        file_metadata = {
            "name":     folder_name,
            "mimeType": "application/vnd.google-apps.folder",
        }
        if parent_id:
            file_metadata["parents"] = [parent_id]
        folder   = service.files().create(
            body=file_metadata, fields="id, name"
        ).execute()
        folder_id = folder["id"]
        logger.info(f"Created folder '{folder_name}' (ID: {folder_id})")
    return folder_id


# ============================================================
# STEP 1: DOWNLOAD CSV FILES FROM GOOGLE DRIVE
# ============================================================

def download_csvs_from_gdrive(service, folder_id, download_dir):
    """Download all CSV files from Google Drive DataSecurity folder."""
    query   = f"'{folder_id}' in parents and name contains '.csv' and trashed=false"
    results = service.files().list(
        q=query,
        fields="files(id, name)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True
    ).execute()

    csv_files = results.get("files", [])
    if not csv_files:
        logger.info("No CSV files found in DataSecurity folder")
        return []

    logger.info(f"Found {len(csv_files)} CSV file(s) in Google Drive")
    downloaded = []

    for file in csv_files:
        file_id   = file["id"]
        file_name = file["name"]
        dest_path = os.path.join(download_dir, file_name)

        try:
            request = service.files().get_media(fileId=file_id)
            with open(dest_path, "wb") as f:
                downloader = MediaIoBaseDownload(f, request)
                done = False
                while not done:
                    _, done = downloader.next_chunk()

            logger.info(f"  Downloaded: {file_name}")
            downloaded.append(dest_path)

        except Exception as e:
            logger.error(f"  Failed to download {file_name}: {e}")
            continue

    logger.info(f"Downloaded {len(downloaded)} file(s)")
    return downloaded


# ============================================================
# STEP 2: PARSE & CONSOLIDATE
# ============================================================

def parse_metadata(filepath):
    """Extract metadata from CSV header rows."""
    meta = {
        "report_name":  None,
        "store_number": None,
        "store_name":   None,
        "report_date":  None,
        "total_count":  0,
    }

    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
    except Exception as e:
        logger.error(f"Cannot read {filepath}: {e}")
        return meta

    if len(lines) < 4:
        return meta

    # Row 2: Report Name
    line2 = lines[1].strip().rstrip(",").strip('"').strip()
    m = re.match(r"Report Name\s*:\s*(.+)", line2)
    if m:
        report_name = m.group(1).strip().strip('"')
        meta["report_name"] = report_name
        parts = re.match(r"(\d+)[.\s]+(.+)", report_name)
        if parts:
            meta["store_number"] = int(parts.group(1))
            meta["store_name"]   = parts.group(2).strip()
        else:
            meta["store_name"] = report_name

    # Row 4: Total Count
    line4 = lines[3].strip().rstrip(",").strip('"').strip()
    m = re.match(r"Total Count\s*:\s*(\d+)", line4)
    if m:
        meta["total_count"] = int(m.group(1))

    # Row 10: Time Frame -> report date
    if len(lines) >= 10:
        line10 = lines[9].strip().rstrip(",").strip('"').strip()
        m = re.search(r"Time Frame\s*:\s*(\w+ \d+,\d{4})", line10)
        if m:
            try:
                meta["report_date"] = datetime.strptime(
                    m.group(1), "%b %d,%Y").strftime("%Y-%m-%d")
            except ValueError:
                pass

    return meta


def read_csv_data(filepath, meta):
    """Read data rows (skip 11 metadata rows)."""
    if meta["total_count"] == 0:
        return None
    try:
        df = pd.read_csv(filepath, skiprows=11, quotechar='"',
                         encoding="utf-8", on_bad_lines="skip")
    except Exception as e:
        logger.error(f"Error reading {filepath}: {e}")
        return None

    if df.empty:
        return None

    df.insert(0, "Store Number", meta["store_number"])
    df.insert(1, "Store Name",   meta["store_name"])
    df["Report Date"] = meta["report_date"]
    df["Source File"] = os.path.basename(filepath)
    return df


def consolidate(file_paths):
    """Read all CSV files and combine into one DataFrame."""
    all_dfs      = []
    empty_stores = []

    for filepath in file_paths:
        fname = os.path.basename(filepath)
        logger.info(f"  Parsing: {fname}")
        meta = parse_metadata(filepath)

        if meta["total_count"] == 0:
            store = meta["report_name"] or fname
            logger.info(f"    No activity: {store}")
            empty_stores.append(store)
            continue

        df = read_csv_data(filepath, meta)
        if df is not None and not df.empty:
            logger.info(f"    Store: {meta['report_name']}, "
                        f"Date: {meta['report_date']}, Rows: {len(df)}")
            all_dfs.append(df)

    if not all_dfs:
        return None, empty_stores

    combined  = pd.concat(all_dfs, ignore_index=True)
    sort_cols = [c for c in ["Store Number", "Store Name",
                              "ACCESSED BY", "TIME MODIFIED"]
                 if c in combined.columns]
    if sort_cols:
        combined = combined.sort_values(by=sort_cols).reset_index(drop=True)

    return combined, empty_stores


# ============================================================
# STEP 3: CREATE EXCEL REPORT
# ============================================================

def create_excel_report(df, output_path):
    """Create a formatted Excel report with teal theme."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Access Report"

    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="007A78")
    hdr_align = Alignment(horizontal="center", vertical="center",
                          wrap_text=True)
    data_font = Font(name="Arial", size=9)
    alt_fill  = PatternFill("solid", fgColor="F0F8F8")
    bdr = Border(
        left=Side(style="thin",   color="CCCCCC"),
        right=Side(style="thin",  color="CCCCCC"),
        top=Side(style="thin",    color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = list(df.columns)

    for col_idx, header in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = bdr

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = data_font
            cell.border = bdr
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=min(100, ws.max_row),
                                 min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max_len + 2

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    logger.info(f"Excel report created: {output_path}")


# ============================================================
# STEP 4a: UPLOAD TO CONSOLIDATED_REPORT (Zapier trigger)
# ============================================================

def upload_report_to_gdrive(service, local_report_path, reports_folder_id):
    """
    Upload the Excel report to Consolidated_Report folder.
    Uses OAuth (your personal account) so no storage quota issues.
    Deletes existing file first so Zapier sees it as brand new.
    """
    file_name = os.path.basename(local_report_path)

    # Delete existing file with same name so Zapier triggers
    query    = (f"name='{file_name}' and '{reports_folder_id}' "
                f"in parents and trashed=false")
    existing = service.files().list(
        q=query, fields="files(id, name)"
    ).execute()
    for f in existing.get("files", []):
        service.files().delete(fileId=f["id"]).execute()
        logger.info(f"  Deleted existing: {f['name']}")

    # Upload new report
    file_metadata = {
        "name":    file_name,
        "parents": [reports_folder_id]
    }
    media = MediaFileUpload(
        local_report_path,
        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        )
    )

    uploaded = service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id, name"
    ).execute()

    logger.info(f"  Uploaded to Consolidated_Report: {uploaded['name']}")
    logger.info(f"  File ID: {uploaded['id']}")
    logger.info("  Zapier will now detect the new file and send the email.")
    return True


# ============================================================
# STEP 4b: SAVE ARCHIVE COPY TO Reports_Archive (permanent)
# ============================================================

def save_archive_copy(service, local_report_path, archive_folder_id):
    """
    Save a permanent timestamped copy to Reports_Archive.
    Never deletes existing files — permanent record.
    """
    timestamp    = datetime.now().strftime("%Y-%m-%d_%H%M")
    base_name    = os.path.splitext(os.path.basename(local_report_path))[0]
    archive_name = f"{base_name}_{timestamp}.xlsx"

    file_metadata = {
        "name":    archive_name,
        "parents": [archive_folder_id]
    }
    media = MediaFileUpload(
        local_report_path,
        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        )
    )

    uploaded = service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id, name"
    ).execute()

    logger.info(f"  Archive copy saved: {uploaded['name']}")
    logger.info(f"  Archive File ID: {uploaded['id']}")
    return True


# ============================================================
# MAIN
# ============================================================

def main():
    logger.info("=" * 60)
    logger.info("DATASECURITY DAILY REPORT")
    logger.info(f"Started: {datetime.now():%Y-%m-%d %H:%M:%S}")
    logger.info("=" * 60)

    # Two separate Drive connections
    logger.info("")
    logger.info("Connecting to Google Drive API")
    logger.info("-" * 40)
    read_service  = get_service_account_service()  # for downloading CSVs
    write_service = get_oauth_service()             # for uploading reports

    # Get source folder (DataSecurity) — via service account
    source_folder_id = get_folder_id(read_service, GDRIVE_SOURCE_FOLDER_NAME)
    if not source_folder_id:
        logger.error("Cannot find DataSecurity folder. Exiting.")
        return

    # Get or create output folders — via OAuth (your account)
    logger.info("")
    logger.info("Setting up output folders")
    logger.info("-" * 40)
    reports_folder_id = get_or_create_folder(
        write_service, GDRIVE_REPORTS_FOLDER_NAME
    )
    archive_folder_id = get_or_create_folder(
        write_service, GDRIVE_ARCHIVE_FOLDER_NAME
    )

    with tempfile.TemporaryDirectory() as tmp_dir:

        # Step 1: Download CSVs
        logger.info("")
        logger.info("STEP 1: Downloading CSV files from Google Drive")
        logger.info("-" * 40)
        downloaded_files = download_csvs_from_gdrive(
            read_service, source_folder_id, tmp_dir
        )
        if not downloaded_files:
            logger.info("No files to process. Exiting.")
            return

        # Step 2: Consolidate
        logger.info("")
        logger.info("STEP 2: Consolidating files")
        logger.info("-" * 40)
        df, empty_stores = consolidate(downloaded_files)
        if df is None or df.empty:
            logger.info("All stores had zero activity. Exiting.")
            return

        store_count  = (df["Store Number"].nunique()
                        if "Store Number" in df.columns else 0)
        record_count = len(df)

        report_date = datetime.now().strftime("%Y-%m-%d")
        if "Report Date" in df.columns and df["Report Date"].notna().any():
            report_date = df["Report Date"].dropna().iloc[0]

        # Step 3: Create Excel report
        logger.info("")
        logger.info("STEP 3: Creating Excel report")
        logger.info("-" * 40)
        output_file = f"DataSecurity_Report_{report_date}.xlsx"
        local_path  = os.path.join(tmp_dir, output_file)
        create_excel_report(df, local_path)
        logger.info(f"  Stores with activity:    {store_count}")
        logger.info(f"  Stores with no activity: {len(empty_stores)}")
        logger.info(f"  Total records:           {record_count}")

        # Step 4a: Upload to Consolidated_Report for Zapier
        logger.info("")
        logger.info("STEP 4a: Uploading to Consolidated_Report (Zapier trigger)")
        logger.info("-" * 40)
        upload_report_to_gdrive(write_service, local_path, reports_folder_id)

        # Step 4b: Save permanent archive copy
        logger.info("")
        logger.info("STEP 4b: Saving archive copy to Reports_Archive")
        logger.info("-" * 40)
        save_archive_copy(write_service, local_path, archive_folder_id)

    logger.info("")
    logger.info("=" * 60)
    logger.info("COMPLETE")
    logger.info("  Zapier will email the report shortly")
    logger.info("  Archive copy saved to: Reports_Archive/")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()